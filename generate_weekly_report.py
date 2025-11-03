"""
SteamWorks Weekly Report Generator
Generates a weekly report for 4 games with key metrics and week-over-week comparisons.
"""

import mysql.connector
from mysql.connector import Error
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
import logging

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('weekly_report_generator.log'),
        logging.StreamHandler()
    ]
)

class WeeklyReportGenerator:
    def __init__(self, db_config):
        self.db_config = db_config
        self.games = [
            {'app_id': 2507950, 'name': 'Delta Force'},
            {'app_id': 2073620, 'name': 'Arena Breakout: Infinite'},
            {'app_id': 3478050, 'name': 'Road to Empress'},
            {'app_id': 3104410, 'name': 'Terminull Brigade'}
        ]
        
    def parse_date_input(self, date_str):
        """Parse YYYYMMDD format to date object"""
        try:
            return datetime.strptime(date_str, '%Y%m%d').date()
        except ValueError:
            raise ValueError(f"Invalid date format: {date_str}. Expected YYYYMMDD (e.g., 20251006)")
    
    def get_date_ranges(self, start_date):
        """Calculate current week and previous week date ranges"""
        # Current week: start_date to start_date + 6 days
        current_week_start = start_date
        current_week_end = start_date + timedelta(days=6)
        
        # Previous week: 7 days before start_date to start_date - 1
        previous_week_start = start_date - timedelta(days=7)
        previous_week_end = start_date - timedelta(days=1)
        
        return {
            'current_week_start': current_week_start,
            'current_week_end': current_week_end,
            'previous_week_start': previous_week_start,
            'previous_week_end': previous_week_end
        }
    
    def fetch_game_data(self, connection, app_id, start_date, end_date):
        """Fetch data for a specific game and date range"""
        cursor = connection.cursor(dictionary=True)
        
        query = """
        SELECT 
            stat_date,
            new_players,
            unique_player,
            daily_total_revenue,
            lifetime_total_revenue,
            dau,
            median_playtime
        FROM game_daily_metrics
        WHERE steam_app_id = %s 
        AND stat_date BETWEEN %s AND %s
        ORDER BY stat_date
        """
        
        cursor.execute(query, (app_id, start_date, end_date))
        results = cursor.fetchall()
        cursor.close()
        
        return results
    
    def validate_data_completeness(self, data, game_name, start_date, end_date):
        """Validate that all 7 days of data exist for a game"""
        expected_days = 7
        actual_days = len(data)
        
        if actual_days != expected_days:
            # Find which dates are missing
            date_range = [start_date + timedelta(days=i) for i in range(7)]
            existing_dates = {row['stat_date'] for row in data}
            missing_dates = [d for d in date_range if d not in existing_dates]
            
            return False, missing_dates
        
        return True, []
    
    def calculate_wow_change(self, current_value, previous_value):
        """Calculate week-over-week percentage change"""
        if previous_value is None or previous_value == 0:
            return None
        
        change = ((current_value - previous_value) / previous_value) * 100
        sign = "+" if change >= 0 else ""
        return f"{sign}{change:.1f}%"
    
    def calculate_metrics(self, current_week_data, previous_week_data):
        """Calculate all metrics for a game"""
        # Sum of new_players
        current_new_players = sum(row['new_players'] or 0 for row in current_week_data)
        previous_new_players = sum(row['new_players'] or 0 for row in previous_week_data)
        new_players_wow = self.calculate_wow_change(current_new_players, previous_new_players)
        
        # Latest unique_player (from last day of current week)
        latest_unique_player = current_week_data[-1]['unique_player']
        
        # Sum of daily_total_revenue
        current_revenue = sum(row['daily_total_revenue'] or 0 for row in current_week_data)
        previous_revenue = sum(row['daily_total_revenue'] or 0 for row in previous_week_data)
        revenue_wow = self.calculate_wow_change(current_revenue, previous_revenue)
        
        # Latest lifetime_total_revenue
        latest_lifetime_revenue = current_week_data[-1]['lifetime_total_revenue']
        
        # Average DAU
        avg_dau = sum(row['dau'] or 0 for row in current_week_data) / len(current_week_data)
        
        # Latest median_playtime
        latest_median_playtime = current_week_data[-1]['median_playtime']
        
        return {
            'sum_new_players': current_new_players,
            'new_players_wow': new_players_wow if new_players_wow else "N/A",
            'latest_unique_player': latest_unique_player,
            'sum_revenue': current_revenue,
            'revenue_wow': revenue_wow if revenue_wow else "N/A",
            'latest_lifetime_revenue': latest_lifetime_revenue,
            'avg_dau': avg_dau,
            'median_playtime': latest_median_playtime or "N/A"
        }
    
    def create_excel_report(self, all_metrics, date_ranges, output_filename):
        """Create formatted Excel report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Report"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Row 1: Game names (merged cells)
        col_idx = 1
        for game in self.games:
            cell = ws.cell(row=1, column=col_idx)
            cell.value = game['name']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            
            # Merge 6 columns for game name
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+5)
            
            # Apply border to all merged cells
            for i in range(6):
                ws.cell(row=1, column=col_idx+i).border = border
            
            col_idx += 6
        
        # Row 2: Column headers
        headers = [
            'Sum New Players\n(WoW %)',
            'Latest\nUnique Player',
            'Sum Revenue\n(WoW %)',
            'Lifetime\nTotal Revenue',
            'Average\nDAU',
            'Median\nPlaytime'
        ]
        
        col_idx = 1
        for game in self.games:
            for header in headers:
                cell = ws.cell(row=2, column=col_idx)
                cell.value = header
                cell.fill = subheader_fill
                cell.font = subheader_font
                cell.alignment = center_alignment
                cell.border = border
                col_idx += 1
        
        # Row 3: Data values
        col_idx = 1
        for game in self.games:
            metrics = all_metrics[game['app_id']]
            
            # Sum New Players (WoW %)
            cell = ws.cell(row=3, column=col_idx)
            cell.value = f"{int(round(metrics['sum_new_players'])):,} ({metrics['new_players_wow']})"
            cell.alignment = center_alignment
            cell.border = border
            col_idx += 1
            
            # Latest Unique Player
            cell = ws.cell(row=3, column=col_idx)
            cell.value = f"{int(round(metrics['latest_unique_player'])):,}" if metrics['latest_unique_player'] else "N/A"
            cell.alignment = center_alignment
            cell.border = border
            col_idx += 1
            
            # Sum Revenue (WoW %)
            cell = ws.cell(row=3, column=col_idx)
            cell.value = f"${int(round(metrics['sum_revenue'])):,} ({metrics['revenue_wow']})"
            cell.alignment = center_alignment
            cell.border = border
            col_idx += 1
            
            # Lifetime Total Revenue
            cell = ws.cell(row=3, column=col_idx)
            cell.value = f"${int(round(metrics['latest_lifetime_revenue'])):,}" if metrics['latest_lifetime_revenue'] else "N/A"
            cell.alignment = center_alignment
            cell.border = border
            col_idx += 1
            
            # Average DAU
            cell = ws.cell(row=3, column=col_idx)
            cell.value = f"{int(round(metrics['avg_dau'])):,}"
            cell.alignment = center_alignment
            cell.border = border
            col_idx += 1
            
            # Median Playtime
            cell = ws.cell(row=3, column=col_idx)
            cell.value = metrics['median_playtime']
            cell.alignment = center_alignment
            cell.border = border
            col_idx += 1
        
        # Auto-adjust column widths
        for col in range(1, 25):  # 24 columns total
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Set row heights
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 25
        
        # Save workbook
        wb.save(output_filename)
        logging.info(f"Excel report saved: {output_filename}")
    
    def generate_report(self, start_date_str):
        """Main method to generate the weekly report"""
        try:
            # Parse input date
            logging.info(f"Parsing start date: {start_date_str}")
            start_date = self.parse_date_input(start_date_str)
            
            # Calculate date ranges
            date_ranges = self.get_date_ranges(start_date)
            logging.info(f"Current week: {date_ranges['current_week_start']} to {date_ranges['current_week_end']}")
            logging.info(f"Previous week: {date_ranges['previous_week_start']} to {date_ranges['previous_week_end']}")
            
            # Connect to database
            logging.info("Connecting to database...")
            connection = mysql.connector.connect(**self.db_config)
            
            # Collect data and validate
            all_metrics = {}
            errors = []
            
            for game in self.games:
                logging.info(f"Processing {game['name']} ({game['app_id']})...")
                
                # Fetch current week data
                current_week_data = self.fetch_game_data(
                    connection,
                    game['app_id'],
                    date_ranges['current_week_start'],
                    date_ranges['current_week_end']
                )
                
                # Validate current week
                is_valid, missing_dates = self.validate_data_completeness(
                    current_week_data,
                    game['name'],
                    date_ranges['current_week_start'],
                    date_ranges['current_week_end']
                )
                
                if not is_valid:
                    missing_str = ", ".join(str(d) for d in missing_dates)
                    errors.append(f"{game['name']}: Missing current week data for dates: {missing_str}")
                    continue
                
                # Fetch previous week data
                previous_week_data = self.fetch_game_data(
                    connection,
                    game['app_id'],
                    date_ranges['previous_week_start'],
                    date_ranges['previous_week_end']
                )
                
                # Validate previous week
                is_valid, missing_dates = self.validate_data_completeness(
                    previous_week_data,
                    game['name'],
                    date_ranges['previous_week_start'],
                    date_ranges['previous_week_end']
                )
                
                if not is_valid:
                    missing_str = ", ".join(str(d) for d in missing_dates)
                    errors.append(f"{game['name']}: Missing previous week data for dates: {missing_str}")
                    continue
                
                # Calculate metrics
                metrics = self.calculate_metrics(current_week_data, previous_week_data)
                all_metrics[game['app_id']] = metrics
                logging.info(f"[OK] {game['name']} metrics calculated successfully")
            
            # Close database connection
            connection.close()
            
            # Check for errors
            if errors:
                logging.error("[ERROR] Data validation failed. Missing data detected:")
                for error in errors:
                    logging.error(f"  - {error}")
                print("\n[ERROR] Cannot generate report due to missing data.")
                print("Missing data details:")
                for error in errors:
                    print(f"  - {error}")
                return False
            
            # Generate Excel report
            output_filename = f"weekly_report_{date_ranges['current_week_start'].strftime('%Y%m%d')}_to_{date_ranges['current_week_end'].strftime('%Y%m%d')}.xlsx"
            logging.info(f"Generating Excel report: {output_filename}")
            self.create_excel_report(all_metrics, date_ranges, output_filename)
            
            print(f"\n[SUCCESS] Weekly report generated successfully!")
            print(f"File: {output_filename}")
            print(f"Period: {date_ranges['current_week_start']} to {date_ranges['current_week_end']}")
            
            return True
            
        except ValueError as e:
            logging.error(f"Input error: {str(e)}")
            print(f"\n[ERROR] {str(e)}")
            return False
        except Error as e:
            logging.error(f"Database error: {str(e)}")
            print(f"\n[ERROR] DATABASE ERROR: {str(e)}")
            return False
        except Exception as e:
            logging.error(f"Unexpected error: {str(e)}")
            print(f"\n[ERROR] UNEXPECTED ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            return False


def main():
    """Main function"""
    print("=" * 60)
    print("SteamWorks Weekly Report Generator")
    print("=" * 60)
    
    # Database configuration
    db_config = {
        'host': 'localhost',
        'port': 3306,
        'database': 'steamworks_crawler',
        'user': 'root',
        'password': 'Zh1149191843!'
    }
    
    # Get user input
    print("\nEnter the starting date (Monday) in YYYYMMDD format")
    print("Example: 20251006 for October 6, 2025")
    start_date_input = input("Start Date: ").strip()
    
    if not start_date_input:
        print("[ERROR] No date provided")
        sys.exit(1)
    
    # Generate report
    generator = WeeklyReportGenerator(db_config)
    success = generator.generate_report(start_date_input)
    
    if success:
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == "__main__":
    main()

